"""XLSX text extractor using openpyxl."""

from __future__ import annotations

from io import BytesIO

try:
    import openpyxl as _openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


def _extract_xlsx_content(wb) -> str:
    """Extract text content from an openpyxl Workbook object."""
    sheets = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = []

        for row in sheet.iter_rows():
            cells = []
            for cell in row:
                if cell.value is not None:
                    cells.append(str(cell.value))
                else:
                    cells.append("")
            if any(c.strip() for c in cells):
                rows.append(" | ".join(cells))

        if rows:
            sheets.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))

    return "\n\n".join(sheets)


class XlsxExtractor:
    """Extract plain text from XLSX bytes."""

    def can_handle(self, doc_type: str) -> bool:
        return doc_type == "xlsx"

    def extract(self, data: bytes, source_path: str | None = None) -> str:
        if not _OPENPYXL_AVAILABLE:
            raise ImportError("pip install chunkeymonkey[xlsx]")

        wb = _openpyxl.load_workbook(BytesIO(data), data_only=True)
        return _extract_xlsx_content(wb)
